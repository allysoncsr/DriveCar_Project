from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Count, Sum, Avg
from datetime import datetime, timedelta
from django.utils import timezone
import json
import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.colors import HexColor
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from .models import Veiculo, RegistroManutencao, Peca

def user_login(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('index')
        else:
            messages.error(request, 'Usuário ou senha inválidos.')
    return render(request, 'drivecar/login.html')

def user_logout(request):
    logout(request)
    return redirect('login')

def register(request):
    if request.method == 'POST':
        username = request.POST['username']
        email = request.POST['email']
        password = request.POST['password']
        password_confirm = request.POST['password_confirm']
        
        if password != password_confirm:
            messages.error(request, 'As senhas não coincidem.')
        elif User.objects.filter(username=username).exists():
            messages.error(request, 'Nome de usuário já existe.')
        elif User.objects.filter(email=email).exists():
            messages.error(request, 'Email já está em uso.')
        elif len(password) < 6:
            messages.error(request, 'A senha deve ter pelo menos 6 caracteres.')
        else:
            user = User.objects.create_user(username=username, email=email, password=password)
            messages.success(request, 'Conta criada com sucesso! Faça login.')
            return redirect('login')
    return render(request, 'drivecar/register.html')

@login_required
def index(request):
    veiculos = Veiculo.objects.filter(usuario=request.user)
    # Simplificar alertas para versão básica
    alertas = []
    
    context = {
        'veiculos': veiculos,
        'alertas': alertas,
        'total_veiculos': veiculos.count(),
        'total_alertas': len(alertas),
    }
    return render(request, 'drivecar/index.html', context)

@login_required
def dashboard(request):
    """Dashboard principal com analytics"""
    veiculos = Veiculo.objects.filter(usuario=request.user)
    registros = RegistroManutencao.objects.filter(veiculo__usuario=request.user)
    
    # Estatísticas gerais
    total_veiculos = veiculos.count()
    total_registros = registros.count()
    gasto_total = registros.aggregate(total=Sum('preco'))['total'] or 0
    km_total = veiculos.aggregate(total=Sum('quilometragem'))['total'] or 0
    
    # Gastos por mês (últimos 6 meses)
    seis_meses_atras = timezone.now() - timedelta(days=180)
    gastos_mensais = []
    meses = []
    
    for i in range(6):
        data_inicial = seis_meses_atras + timedelta(days=30*i)
        data_final = data_inicial + timedelta(days=30)
        gasto_mes = registros.filter(
            data__gte=data_inicial,
            data__lt=data_final
        ).aggregate(total=Sum('preco'))['total'] or 0
        
        gastos_mensais.append(float(gasto_mes))
        meses.append(data_inicial.strftime('%b/%Y'))
    
    # Distribuição por tipo de peça
    tipos_peca = registros.values('peca__nome').annotate(
        total=Sum('preco'),
        quantidade=Count('id')
    ).order_by('-total')[:5]
    
    context = {
        'total_veiculos': total_veiculos,
        'total_registros': total_registros,
        'gasto_total': gasto_total,
        'km_total': km_total,
        'alertas_ativos': 0,  # Sistema de alertas desabilitado temporariamente
        'gastos_mensais': json.dumps(gastos_mensais),
        'meses': json.dumps(meses),
        'tipos_peca': tipos_peca,
        'veiculos': veiculos,
    }
    return render(request, 'drivecar/dashboard.html', context)

@login_required
def dashboard_veiculo(request, veiculo_id):
    """Dashboard específico de um veículo"""
    veiculo = get_object_or_404(Veiculo, id=veiculo_id, usuario=request.user)
    registros = RegistroManutencao.objects.filter(veiculo=veiculo)
    
    # Estatísticas do veículo
    total_registros = registros.count()
    gasto_total = registros.aggregate(total=Sum('preco'))['total'] or 0
    gasto_medio = registros.aggregate(media=Avg('preco'))['media'] or 0
    
    # Gastos por mês (últimos 12 meses)
    doze_meses_atras = timezone.now() - timedelta(days=365)
    gastos_mensais = []
    meses = []
    
    for i in range(12):
        data_inicial = doze_meses_atras + timedelta(days=30*i)
        data_final = data_inicial + timedelta(days=30)
        gasto_mes = registros.filter(
            data__gte=data_inicial,
            data__lt=data_final
        ).aggregate(total=Sum('preco'))['total'] or 0
        
        gastos_mensais.append(float(gasto_mes))
        meses.append(data_inicial.strftime('%b/%Y'))
    
    # Peças mais trocadas
    pecas_frequentes = registros.values('peca__nome').annotate(
        quantidade=Count('id'),
        gasto_total=Sum('preco')
    ).order_by('-quantidade')
    
    context = {
        'veiculo': veiculo,
        'total_registros': total_registros,
        'gasto_total': gasto_total,
        'gasto_medio': gasto_medio,
        'gastos_mensais': json.dumps(gastos_mensais),
        'meses': json.dumps(meses),
        'pecas_frequentes': pecas_frequentes,
        'proximas_manutencoes': [],  # Sistema de alertas desabilitado
        'registros_recentes': registros.order_by('-data')[:5],
    }
    return render(request, 'drivecar/dashboard_veiculo.html', context)

@login_required
def export_pdf(request):
    """Exportar relatório em PDF"""
    # Criar buffer
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    story = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        spaceAfter=30,
        textColor=colors.HexColor('#2c3e50'),
        alignment=1  # Center
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=12,
        textColor=colors.HexColor('#34495e'),
    )
    
    # Título
    title = Paragraph("Relatório DriveCar", title_style)
    story.append(title)
    story.append(Spacer(1, 20))
    
    # Dados do usuário
    user_info = Paragraph(f"<b>Usuário:</b> {request.user.username}<br/><b>Data:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal'])
    story.append(user_info)
    story.append(Spacer(1, 20))
    
    # Veículos
    veiculos = Veiculo.objects.filter(usuario=request.user)
    story.append(Paragraph("Seus Veículos", heading_style))
    
    if veiculos:
        veiculo_data = [['Marca/Modelo', 'Ano', 'Combustível', 'KM', 'Cor']]
        for veiculo in veiculos:
            veiculo_data.append([
                f"{veiculo.marca} {veiculo.modelo}",
                str(veiculo.ano),
                veiculo.combustivel,
                f"{veiculo.quilometragem:,}".replace(',', '.'),
                veiculo.cor
            ])
        
        veiculo_table = Table(veiculo_data, colWidths=[2.5*inch, 0.8*inch, 1*inch, 1*inch, 1*inch])
        veiculo_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(veiculo_table)
    else:
        story.append(Paragraph("Nenhum veículo cadastrado.", styles['Normal']))
    
    story.append(Spacer(1, 20))
    
    # Registros de Manutenção
    registros = RegistroManutencao.objects.filter(veiculo__usuario=request.user).order_by('-data')[:10]
    story.append(Paragraph("Últimas Manutenções", heading_style))
    
    if registros:
        registro_data = [['Data', 'Veículo', 'Peça', 'KM', 'Custo']]
        for registro in registros:
            registro_data.append([
                registro.data.strftime('%d/%m/%Y'),
                f"{registro.veiculo.marca} {registro.veiculo.modelo}",
                registro.peca.nome,
                f"{registro.km:,}".replace(',', '.'),
                f"R$ {registro.preco:,.2f}".replace(',', '.')
            ])
        
        registro_table = Table(registro_data, colWidths=[1*inch, 2*inch, 1.5*inch, 1*inch, 1*inch])
        registro_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e74c3c')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(registro_table)
    else:
        story.append(Paragraph("Nenhuma manutenção registrada.", styles['Normal']))
    
    # Construir PDF
    doc.build(story)
    buffer.seek(0)
    
    # Resposta HTTP
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="drivecar_relatorio_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    return response

@login_required
def export_excel(request):
    """Exportar dados em Excel"""
    # Criar workbook
    wb = openpyxl.Workbook()
    
    # Aba de Veículos
    ws_veiculos = wb.active
    ws_veiculos.title = "Veículos"
    
    # Headers
    headers_veiculos = ['Marca', 'Modelo', 'Ano', 'Cor', 'Combustível', 'Quilometragem', 'Data Cadastro']
    ws_veiculos.append(headers_veiculos)
    
    # Estilo do header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col_num, header in enumerate(headers_veiculos, 1):
        cell = ws_veiculos.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Dados dos veículos
    veiculos = Veiculo.objects.filter(usuario=request.user)
    for veiculo in veiculos:
        ws_veiculos.append([
            veiculo.marca,
            veiculo.modelo,
            veiculo.ano,
            veiculo.cor,
            veiculo.combustivel,
            veiculo.quilometragem,
            veiculo.data_cadastro.strftime('%d/%m/%Y') if veiculo.data_cadastro else ''
        ])
    
    # Ajustar largura das colunas
    for column in ws_veiculos.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_veiculos.column_dimensions[column_letter].width = adjusted_width
    
    # Aba de Manutenções
    ws_manutencoes = wb.create_sheet(title="Manutenções")
    
    headers_manutencoes = ['Data', 'Veículo', 'Peça', 'Quilometragem', 'Custo', 'Troca', 'Garantia (meses)', 'Observações']
    ws_manutencoes.append(headers_manutencoes)
    
    # Estilo do header
    for col_num, header in enumerate(headers_manutencoes, 1):
        cell = ws_manutencoes.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        cell.alignment = header_alignment
    
    # Dados das manutenções
    registros = RegistroManutencao.objects.filter(veiculo__usuario=request.user).order_by('-data')
    for registro in registros:
        ws_manutencoes.append([
            registro.data.strftime('%d/%m/%Y'),
            f"{registro.veiculo.marca} {registro.veiculo.modelo}",
            registro.peca.nome,
            registro.km,
            float(registro.preco),
            'Sim' if registro.troca else 'Não',
            registro.garantia_meses or '',
            registro.observacoes or ''
        ])
    
    # Ajustar largura das colunas
    for column in ws_manutencoes.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_manutencoes.column_dimensions[column_letter].width = adjusted_width
    
    # Salvar em buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Resposta HTTP
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="drivecar_dados_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    return response

@login_required
def offline(request):
    """Página offline para PWA"""
    return render(request, 'drivecar/offline.html')

@login_required
def cadastrar_veiculo(request):
    if request.method == 'POST':
        marca = request.POST['marca']
        modelo = request.POST['modelo']
        ano = int(request.POST['ano'])
        cor = request.POST['cor']
        combustivel = request.POST['combustivel']
        quilometragem = int(request.POST['quilometragem'])
        
        veiculo = Veiculo.objects.create(
            usuario=request.user,
            marca=marca,
            modelo=modelo,
            ano=ano,
            cor=cor,
            combustivel=combustivel,
            quilometragem=quilometragem
        )
        
        messages.success(request, f'Veículo {marca} {modelo} cadastrado com sucesso!')
        return redirect('index')
    
    return render(request, 'drivecar/cadastrar_veiculo.html')

@login_required
def editar_veiculo(request, veiculo_id):
    veiculo = get_object_or_404(Veiculo, id=veiculo_id, usuario=request.user)
    
    if request.method == 'POST':
        veiculo.marca = request.POST['marca']
        veiculo.modelo = request.POST['modelo']
        veiculo.ano = int(request.POST['ano'])
        veiculo.cor = request.POST['cor']
        veiculo.combustivel = request.POST['combustivel']
        veiculo.quilometragem = int(request.POST['quilometragem'])
        veiculo.save()
        
        messages.success(request, 'Veículo atualizado com sucesso!')
        return redirect('index')
    
    return render(request, 'drivecar/editar_veiculo.html', {'veiculo': veiculo})

@login_required
def excluir_veiculo(request, veiculo_id):
    veiculo = get_object_or_404(Veiculo, id=veiculo_id, usuario=request.user)
    
    if request.method == 'POST':
        veiculo.delete()
        messages.success(request, 'Veículo excluído com sucesso!')
        return redirect('index')
    
    return render(request, 'drivecar/confirmar_exclusao.html', {'veiculo': veiculo})

@login_required
def manutencao(request, veiculo_id):
    veiculo = get_object_or_404(Veiculo, id=veiculo_id, usuario=request.user)
    pecas = Peca.objects.all()
    registros = RegistroManutencao.objects.filter(veiculo=veiculo).order_by('-data')
    
    if request.method == 'POST':
        peca_id = request.POST['peca']
        quilometragem = int(request.POST['quilometragem'])
        custo = float(request.POST['custo'].replace(',', '.'))
        observacoes = request.POST.get('observacoes', '')
        data_realizacao = request.POST['data_realizacao']
        troca = request.POST.get('troca') == 'on'
        garantia_meses = request.POST.get('garantia_meses')
        
        # Obter peça
        peca = get_object_or_404(Peca, id=peca_id)
        
        registro = RegistroManutencao.objects.create(
            veiculo=veiculo,
            peca=peca,
            km=quilometragem,
            preco=custo,
            observacoes=observacoes,
            data=datetime.strptime(data_realizacao, '%Y-%m-%d').date(),
            troca=troca,
            garantia_meses=int(garantia_meses) if garantia_meses else None,
            alerta_ativo=False  # Sistema de alertas foi removido
        )
        
        # Atualizar quilometragem do veículo se necessário
        if quilometragem > veiculo.quilometragem:
            veiculo.quilometragem = quilometragem
            veiculo.save()
        
        messages.success(request, 'Registro de manutenção adicionado com sucesso!')
        return redirect('manutencao', veiculo_id=veiculo.id)
    
    context = {
        'veiculo': veiculo,
        'pecas': pecas,
        'registros': registros,
    }
    return render(request, 'drivecar/manutencao.html', context)

@login_required
def alertas(request):
    # Sistema de alertas temporariamente desabilitado
    alertas = []
    
    context = {
        'alertas': alertas,
        'total_alertas': 0,
    }
    return render(request, 'drivecar/alertas.html', context)

@login_required
def alertas_veiculo_fragment(request, veiculo_id):
    """Fragment para atualizar alertas via HTMX"""
    veiculo = get_object_or_404(Veiculo, id=veiculo_id, usuario=request.user)
    alertas = []  # Sistema de alertas desabilitado
    
    context = {
        'veiculo': veiculo,
        'alertas': alertas,
    }
    return render(request, 'drivecar/alertas_veiculo_fragment.html', context)

@login_required
def registros_fragment(request, veiculo_id):
    """Fragment para atualizar registros via HTMX"""
    veiculo = get_object_or_404(Veiculo, id=veiculo_id, usuario=request.user)
    registros = RegistroManutencao.objects.filter(veiculo=veiculo).order_by('-data')
    
    context = {
        'veiculo': veiculo,
        'registros': registros,
    }
    return render(request, 'drivecar/registros_fragment.html', context)

@login_required
def editar_registro(request, registro_id):
    registro = get_object_or_404(RegistroManutencao, id=registro_id, veiculo__usuario=request.user)
    
    if request.method == 'POST':
        peca_id = request.POST['peca']
        quilometragem = int(request.POST['quilometragem'])
        custo = float(request.POST['custo'].replace(',', '.'))
        observacoes = request.POST.get('observacoes', '')
        data_realizacao = request.POST['data_realizacao']
        troca = request.POST.get('troca') == 'on'
        garantia_meses = request.POST.get('garantia_meses')
        
        # Atualizar registro
        peca = get_object_or_404(Peca, id=peca_id)
        registro.peca = peca
        registro.km = quilometragem
        registro.preco = custo
        registro.observacoes = observacoes
        registro.data = datetime.strptime(data_realizacao, '%Y-%m-%d').date()
        registro.troca = troca
        registro.garantia_meses = int(garantia_meses) if garantia_meses else None
        
        registro.save()
        
        messages.success(request, 'Registro atualizado com sucesso!')
        return redirect('manutencao', veiculo_id=registro.veiculo.id)
    
    pecas = Peca.objects.all()
    context = {
        'registro': registro,
        'pecas': pecas,
    }
    return render(request, 'drivecar/editar_registro.html', context)

@login_required
def excluir_registro(request, registro_id):
    registro = get_object_or_404(RegistroManutencao, id=registro_id, veiculo__usuario=request.user)
    veiculo_id = registro.veiculo.id
    
    if request.method == 'POST':
        registro.delete()
        messages.success(request, 'Registro excluído com sucesso!')
        return redirect('manutencao', veiculo_id=veiculo_id)
    
    return render(request, 'drivecar/confirmar_exclusao_registro.html', {'registro': registro})

@login_required
def desativar_alerta(request, registro_id):
    """Desativar alerta de um registro específico"""
    if request.method == 'POST':
        registro = get_object_or_404(RegistroManutencao, id=registro_id, veiculo__usuario=request.user)
        registro.alerta_ativo = False
        registro.save()
        
        if request.headers.get('HX-Request'):
            return JsonResponse({'success': True, 'message': 'Alerta desativado com sucesso!'})
        else:
            messages.success(request, 'Alerta desativado com sucesso!')
            return redirect('alertas')
    
    return JsonResponse({'success': False, 'message': 'Método não permitido'})

@login_required
def lista_alertas(request):
    """Lista de alertas em JSON para API"""
    alertas = []  # Sistema de alertas desabilitado
    return JsonResponse({'alertas': alertas, 'total': 0})